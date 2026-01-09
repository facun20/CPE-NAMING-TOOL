import streamlit as st
import anthropic
import requests
import base64
import re
import json
from datetime import datetime
from io import BytesIO

# OpenRouter API configuration (users provide their own API key)
OPENROUTER_BASE_URL = "https://openrouter.ai/api/v1/chat/completions"
GEMINI_MODEL = "google/gemini-2.5-flash"

# PDF parsing and rendering
try:
    import fitz  # PyMuPDF - for PDF to image conversion (OCR support)
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    from PyPDF2 import PdfReader
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

# Word document parsing
try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

# Excel parsing
try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

# Image handling
try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# Page configuration
st.set_page_config(
    page_title="UBC CPE File Naming Tool",
    page_icon="📁",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for UBC branding
st.markdown("""
<style>
    :root {
        --ubc-blue: #002145;
        --ubc-gold: #C1A01E;
    }

    .main-header {
        text-align: center;
        padding: 20px;
        border-bottom: 3px solid #002145;
        margin-bottom: 20px;
    }

    .main-header h1 {
        color: #002145;
        margin-bottom: 5px;
    }

    .version-badge {
        background-color: #C1A01E;
        color: white;
        padding: 4px 12px;
        border-radius: 12px;
        font-size: 14px;
        font-weight: 600;
    }

    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }

    .stTabs [data-baseweb="tab"] {
        background-color: #f0f0f0;
        border-radius: 8px 8px 0 0;
        padding: 10px 20px;
    }

    .stTabs [aria-selected="true"] {
        background-color: #002145 !important;
        color: white !important;
    }

    .output-box {
        background-color: #f8f8f8;
        border-left: 4px solid #C1A01E;
        padding: 15px;
        border-radius: 0 8px 8px 0;
        margin: 10px 0;
    }

    .file-suggestion {
        background-color: #e8f4f8;
        border: 1px solid #002145;
        border-radius: 8px;
        padding: 15px;
        margin: 10px 0;
    }

    .help-panel {
        background-color: #f8f8f8;
        border-left: 4px solid #C1A01E;
        padding: 15px;
        border-radius: 0 8px 8px 0;
    }
</style>
""", unsafe_allow_html=True)

# Document form codes
DOCUMENT_FORMS = {
    "": "Select document form",
    "AGD": "AGD - Agenda",
    "AGR": "AGR - Agreement",
    "ANN": "ANN - Announcement",
    "APP": "APP - Appendix",
    "ATD": "ATD - Attendance",
    "BGT": "BGT - Course Budget",
    "BRN": "BRN - Briefing Note",
    "CCP": "CCP - Concept Paper",
    "CON": "CON - Contract",
    "DAT": "DAT - Data Set",
    "FCT": "FCT - Fact Sheet",
    "FRM": "FRM - Form",
    "GRA": "GRA - Grant",
    "GUI": "GUI - Guidelines",
    "INS": "INS - Instruction",
    "INT": "INT - Interview",
    "INV": "INV - Invoice",
    "LAT": "LAT - Letter of Attendance",
    "LCO": "LCO - Letter of Completion",
    "LGL": "LGL - Legal Document",
    "LPA": "LPA - Letter of Participation",
    "LPR": "LPR - Letter of Proficiency",
    "LTR": "LTR - Letter",
    "MIN": "MIN - Minutes",
    "MNL": "MNL - Manual",
    "NCC": "NCC - Non-Credit Certificate",
    "NCM": "NCM - Non-Credit MicroCertificate",
    "PLN": "PLN - Plan",
    "POL": "POL - Policy",
    "PRC": "PRC - Procedure",
    "PRO": "PRO - Proposal",
    "PRS": "PRS - Presentation",
    "PST": "PST - Poster",
    "RPT": "RPT - Report",
    "RVW": "RVW - Review",
    "SCH": "SCH - Schedule",
    "SMP": "SMP - Sample",
    "SRY": "SRY - Survey",
    "SUM": "SUM - Summary",
    "TEM": "TEM - Template",
    "TML": "TML - Timeline",
    "IMG": "IMG - Image",
    "SCR": "SCR - Screenshot",
    "DIA": "DIA - Diagram"
}

REVISION_STATUSES = {
    "A": "A - Initial draft sent for review",
    "B": "B - Official draft sent for external or internal review",
    "C": "C - Next incarnation of official draft",
    "0": "0 - First final revision",
    "0A": "0A - Draft after final has been produced",
    "0B": "0B - Draft after final has been produced",
    "0C": "0C - Draft after final has been produced",
    "1": "1 - Next revision after final"
}

FILE_EXTENSIONS = {
    "pdf": "pdf - Portable Document Format",
    "docx": "docx - Word Document",
    "xlsx": "xlsx - Excel Spreadsheet",
    "pptx": "pptx - PowerPoint Presentation",
    "txt": "txt - Text File",
    "csv": "csv - Comma Separated Values",
    "jpg": "jpg - JPEG Image",
    "png": "png - PNG Image"
}

HELP_CONTENT = {
    "subject": {
        "title": "Subject/Activity (Required)",
        "content": "The main topic or activity the document covers. Use PascalCase formatting (capitalize the first letter of each word with no spaces). Examples: NamingConventions, RecordsManagement, CourseEvaluation"
    },
    "date": {
        "title": "Date (Required)",
        "content": "The date when the document was created or last modified. Will be formatted as YYYY-MM-DD according to ISO 8601 standard. Example: 2025-05-28"
    },
    "revisionStatus": {
        "title": "Revision Status (Required)",
        "content": "Indicates the version and status: Letters (A, B, C) for drafts, Numbers (0, 1, 2) for final versions, Combinations (0A, 0B) for drafts after final version"
    },
    "projectCode": {
        "title": "Project/Account Number (Optional)",
        "content": "A control number, project code, or account identifier that helps organize documents. Examples: CPE, PROJ2024, ACC-001"
    },
    "documentForm": {
        "title": "Document Form (Optional)",
        "content": "Three-letter code indicating the type of document. Common forms: AGD (Agenda), RPT (Report), GUI (Guidelines), PRS (Presentation), MNL (Manual)"
    },
    "facultySchool": {
        "title": "Faculty-School (Course Format)",
        "content": "Identifies the faculty and school using a dash separator. Examples: FHSD-SoN (Faculty of Health & Social Development - School of Nursing)"
    },
    "courseCode": {
        "title": "Course Code (Course Format)",
        "content": "Four-digit course number followed by four-digit section code, separated by a dash. Format: ####-#### Examples: 0386-0001"
    },
    "termOffered": {
        "title": "Term Offered (Course Format)",
        "content": "Academic term using format YYYYST where YYYY = Year, S = Session (W=Winter, S=Summer), T = Term (1, 2). Examples: 2024WT1, 2025ST1"
    }
}


def pdf_to_image(file_bytes: bytes) -> str:
    """Convert first page of PDF to base64 image for AI vision/OCR.

    Returns base64 data URI string for image, or None if conversion fails.
    """
    if not PYMUPDF_AVAILABLE:
        return None

    try:
        # Open PDF from bytes
        pdf_doc = fitz.open(stream=file_bytes, filetype="pdf")

        # Get first page (most important for document identification)
        page = pdf_doc[0]

        # Render page to image at 150 DPI (good balance of quality vs size)
        mat = fitz.Matrix(150/72, 150/72)  # 150 DPI
        pix = page.get_pixmap(matrix=mat)

        # Convert to PNG bytes
        img_bytes = pix.tobytes("png")

        # Encode as base64
        base64_data = base64.b64encode(img_bytes).decode('utf-8')

        pdf_doc.close()

        return f"data:image/png;base64,{base64_data}"
    except Exception as e:
        return None


def read_pdf_content(file_bytes: bytes) -> str:
    """Extract text from PDF file (fallback when image conversion not available)."""
    if not PDF_AVAILABLE:
        return "PDF parsing not available"

    try:
        pdf_reader = PdfReader(BytesIO(file_bytes))
        text = ""
        for page in pdf_reader.pages[:10]:  # Limit to first 10 pages
            text += page.extract_text() or ""
        return text[:4000]
    except Exception as e:
        return f"Could not extract PDF text: {str(e)}"


def read_docx_content(file_bytes: bytes) -> str:
    """Extract text from Word document."""
    if not DOCX_AVAILABLE:
        return "Word document parsing not available"

    try:
        doc = Document(BytesIO(file_bytes))
        text = "\n".join([para.text for para in doc.paragraphs])
        return text[:4000]
    except Exception as e:
        return f"Could not extract Word text: {str(e)}"


def read_xlsx_content(file_bytes: bytes) -> str:
    """Extract text from Excel file with better structure preservation."""
    if not XLSX_AVAILABLE:
        return "Excel parsing not available"

    try:
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        text = "=== EXCEL FILE CONTENT ===\n"

        for sheet_name in workbook.sheetnames[:3]:  # Limit to first 3 sheets
            sheet = workbook[sheet_name]
            text += f"\n--- Sheet: {sheet_name} ---\n"

            # Get actual used range
            max_row = min(sheet.max_row or 1, 100)  # Limit to 100 rows
            max_col = min(sheet.max_column or 1, 20)  # Limit to 20 columns

            row_count = 0
            for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                # Skip completely empty rows
                if any(cell for cell in row):
                    row_text = " | ".join([str(cell)[:50] if cell else "" for cell in row])
                    text += row_text.strip() + "\n"
                    row_count += 1
                    if row_count >= 75:  # Limit rows per sheet
                        text += "... (more rows)\n"
                        break

        return text[:6000]  # Allow more content for Excel
    except Exception as e:
        return f"Could not extract Excel text: {str(e)}"


def read_file_content(uploaded_file) -> tuple:
    """Read content from uploaded file. Returns (content, content_type).

    For PDFs: Converts to image for AI vision/OCR (much better text detection).
    Falls back to text extraction if image conversion unavailable.
    """
    file_bytes = uploaded_file.read()
    file_name = uploaded_file.name.lower()

    if file_name.endswith('.pdf'):
        # Try to convert PDF to image for AI vision/OCR (best accuracy)
        pdf_image = pdf_to_image(file_bytes)
        if pdf_image:
            return pdf_image, "image"
        # Fallback to text extraction if PyMuPDF not available
        return read_pdf_content(file_bytes), "text"
    elif file_name.endswith('.docx'):
        return read_docx_content(file_bytes), "text"
    elif file_name.endswith(('.xlsx', '.xls')):
        return read_xlsx_content(file_bytes), "text"
    elif file_name.endswith(('.txt', '.csv')):
        return file_bytes.decode('utf-8', errors='ignore')[:4000], "text"
    elif file_name.endswith(('.jpg', '.jpeg', '.png', '.gif')):
        # For images, return base64 encoded data
        ext = file_name.split('.')[-1]
        mime_type = f"image/{ext}" if ext != 'jpg' else "image/jpeg"
        base64_data = base64.b64encode(file_bytes).decode('utf-8')
        return f"data:{mime_type};base64,{base64_data}", "image"
    else:
        return "Unsupported file type", "unknown"


def analyze_with_claude(api_key: str, content: str, file_name: str, content_type: str, privacy_level: str) -> dict:
    """Analyze file content with Claude API."""
    try:
        client = anthropic.Anthropic(api_key=api_key)
        today = datetime.now().strftime("%Y-%m-%d")

        if content_type == "image":
            # Image analysis
            mime_type = content.split(';')[0].replace('data:', '')
            base64_data = content.split(',')[1]

            # Determine if this is likely a PDF (rendered as image) or actual image
            is_pdf = file_name.lower().endswith('.pdf')
            ext = "pdf" if is_pdf else file_name.split('.')[-1] if '.' in file_name else "jpg"

            if is_pdf:
                # PDF rendered as image - Claude analysis with OCR
                image_prompt = f"""You are a file naming expert for UBC CPE (Continuing Professional Education). READ ALL TEXT in this PDF document image carefully.

Current filename: {file_name}
Today's date: {today}

=== STEP 1: EXTRACT ALL AVAILABLE FIELDS ===

1. FACULTY-SCHOOL (use dash between Faculty and School):
   UBC Okanagan Faculties & Schools - use these abbreviations:
   - "Irving K. Barber Faculty of Arts and Social Sciences" → IKBASS
   - "Irving K. Barber Faculty of Science" → IKBFOS (or IKB-FOS if with school)
   - "Faculty of Creative and Critical Studies" → FCCS
   - "Okanagan School of Education" → OSE
   - "Faculty of Applied Science" / "School of Engineering" → APSC-SoE
   - "Faculty of Health and Social Development" → FHSD
     - with "School of Nursing" → FHSD-SoN
     - with "School of Social Work" → FHSD-SSW
     - with "School of Health and Exercise Sciences" → FHSD-SHES
   - "Faculty of Management" → FoM
   - "Faculty of Medicine" → MED
   - "College of Graduate Studies" → CoGS

   FORMAT: Faculty-School with dash (e.g., FHSD-SoN, APSC-SoE)
   If only Faculty found (no specific school), just use Faculty code (e.g., FCCS)

2. COURSE CODE - Four digits + dash + four digits (e.g., 0386-0001)
   Look for course numbers, section codes

3. TERM OFFERED - Format: YYYYST (Year + Session + Term)
   - Session: W = Winter (Sept-Apr), S = Summer (May-Aug)
   - Term: 1 or 2
   - Sept-Dec → WT1 | Jan-Apr → WT2 | May-Jun → ST1 | Jul-Aug → ST2
   - Example: "December 2025" → 2025WT1
   - Example: "March 2026" → 2025WT2 (still Winter session)

4. PROJECT/ACCOUNT CODE - Project numbers, grant codes (e.g., CPE, PROJ2024)

5. SUBJECT - Use Title-Kebab-Case (e.g., Wildland-Fire-Ecology)
   ⚠️ MAX 50 CHARACTERS! Truncate at word boundary if longer.
   Example: "Foundations-For-A-Restorative-Approach-Health-Care" (50 chars max)

6. DOCUMENT FORM codes:
   LETTERS & CERTIFICATES:
   - "Letter of Proficiency" → LPR | "Letter of Completion" → LCO
   - "Letter of Attendance" → LAT | "Letter of Participation" → LPA
   - "Non-Credit Certificate" → NCC | "MicroCertificate" → NCM

   COMMON TYPES:
   - Agenda → AGD | Agreement → AGR | Budget → BGT | Contract → CON
   - Form → FRM | Guidelines/Guide → GUI | Instructions → INS | Invoice → INV
   - Letter (general) → LTR | Minutes → MIN | Manual → MNL | Plan → PLN
   - Policy → POL | Procedure/SOP → PRC | Proposal → PRO | Presentation → PRS
   - Report → RPT | Schedule → SCH | Template → TEM | Summary → SUM

7. REVISION STATUS:
   - Draft: RevA, RevB, RevC (letters for drafts)
   - Final: Rev0 (first final), Rev1, Rev2 (subsequent finals)
   - Draft after final: Rev0A, Rev0B (number + letter)

=== STEP 2: CHOOSE FORMAT ===

COURSE FORMAT (Faculty + Course content):
Faculty-School_CourseCode_TermOffered_DocumentForm_Date_RevisionStatus.ext
Example: FHSD-SoN_0386-0001_2024WT2_TEM_2025-01-10_Rev0.pptx

ADVANCED FORMAT (Project code found):
ProjectCode_Subject_DocumentForm_Date_RevisionStatus.ext
Example: CPE_Records-Management_POL_2025-01-20_Rev0.pdf

BASIC FORMAT (minimal info):
Subject_DocumentForm_Date_RevisionStatus.ext
Example: Naming-Conventions_LPR_2025-03-11_RevA.docx

=== STEP 3: RESPOND WITH JSON (no markdown) ===
{{"suggestedName": "GeneratedFilename.pdf", "formatUsed": "course|advanced|basic", "extractedFields": {{"facultySchool": "FHSD-SoN or null", "courseCode": "0386-0001 or null", "term": "2024WT2 or null", "projectCode": "CPE or null", "subject": "Subject-In-Title-Kebab-Case", "documentForm": "CODE", "date": "YYYY-MM-DD", "revision": "Rev0"}}, "reasoning": "I found [text]. Faculty: [X]. School: [X]. Term: [X]. Document type: [X].", "confidence": 9}}"""
            else:
                # Actual image file (photo, screenshot, diagram)
                image_prompt = f"""I need help creating a CPE-compliant filename for this image.

IMPORTANT: First, carefully analyze what is shown in this image. Describe what you see in detail.

Current filename: {file_name}

Key elements for the filename:
- Subject (required): What is actually shown in the image. Use Title-Kebab-Case (e.g., Campus-Building). MAX 50 CHARACTERS.
- Document Form: Use IMG for photos, SCR for screenshots, DIA for diagrams
- Date: Use today's date {today} if no date is visible
- Revision: Use 'Rev0' for final version

Respond with ONLY a JSON object:
{{"suggestedName": "Content-Description_IMG_{today}_Rev0.{ext}", "reasoning": "I can see [description]. I chose [Subject] because [reason]. I used IMG/SCR/DIA because [reason].", "confidence": 8, "detectedType": "IMG", "suggestedSubject": "Subject-In-Title-Kebab-Case"}}"""

            message = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=1000,
                messages=[{
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": image_prompt
                        },
                        {
                            "type": "image",
                            "source": {
                                "type": "base64",
                                "media_type": mime_type,
                                "data": base64_data
                            }
                        }
                    ]
                }]
            )
        else:
            # Text document analysis - send full content for best analysis
            document_content = content[:6000]  # Allow more for Excel/text files

            message = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=1000,
                messages=[{
                    "role": "user",
                    "content": f"""You are a file naming expert for UBC CPE (Continuing Professional Education). Analyze this document and suggest a CPE-compliant filename.

Current file: {file_name}
Today's date: {today}

=== STEP 1: CHECK FOR FACULTY-SCHOOL (STRICT MATCHING ONLY) ===
ONLY use Faculty-School codes if you find EXACT matches to these UBC Okanagan names:
- "Irving K. Barber Faculty of Arts and Social Sciences" → IKBASS
- "Irving K. Barber Faculty of Science" → IKBFOS
- "Faculty of Creative and Critical Studies" → FCCS
- "Okanagan School of Education" → OSE
- "Faculty of Applied Science" + "School of Engineering" → APSC-SoE
- "Faculty of Health and Social Development" → FHSD
  - + "School of Nursing" → FHSD-SoN
  - + "School of Social Work" → FHSD-SSW
- "Faculty of Management" → FoM
- "Faculty of Medicine" → MED
- "College of Graduate Studies" → CoGS

⚠️ DO NOT invent Faculty-School codes! If you don't see these EXACT faculty names, set facultySchool to null.
⚠️ "CPE", "ITAL", "55PLUS", course abbreviations, etc. are NOT Faculty-School codes!

=== STEP 2: EXTRACT OTHER FIELDS ===
- COURSE CODE: Four digits + dash + four digits (e.g., 0386-0001). Must be this exact format.
- TERM: Format YYYYST (e.g., 2025WT1). Only if you see clear term/session info.
- PROJECT CODE: Only if explicitly labeled (e.g., "Project: CPE", "Account: XYZ")
- SUBJECT: Use Title-Kebab-Case (e.g., Foundations-For-Restorative-Approach)
  ⚠️ MAX 50 CHARACTERS for subject! Truncate at word boundary if longer.
  Example: "Foundations For A Restorative Approach Health Care Harm And Wellbeing"
  → Truncate to: "Foundations-For-A-Restorative-Approach-Health-Care" (50 chars)
- DOCUMENT FORM: Match keywords to codes:
  - Report/Enrollment/Count/Chart → RPT | Survey/Analysis → SRY | Summary → SUM
  - Letter of Proficiency → LPR | Letter of Completion → LCO
  - Guidelines/Guide → GUI | Manual → MNL | Template → TEM | Schedule → SCH
  - Budget → BGT | Invoice → INV | Contract → CON | Form → FRM
- REVISION: RevA/B/C for drafts, Rev0 for first final, Rev1/2 for subsequent

=== STEP 3: CHOOSE FORMAT ===
COURSE FORMAT (ONLY if you found valid Faculty-School AND course content):
Faculty-School_CourseCode_Term_DocumentForm_Date_Revision.ext

ADVANCED FORMAT (if project code is explicitly present):
ProjectCode_Subject_DocumentForm_Date_Revision.ext

BASIC FORMAT (default - use this for most files):
Subject_DocumentForm_Date_Revision.ext
Example: Enrollment-Count-Report_RPT_{today}_Rev0.xls

=== STEP 4: RESPOND WITH JSON ===
Content: {document_content}

Respond with ONLY this JSON (no markdown):
{{"suggestedName": "Filename.ext", "formatUsed": "course|advanced|basic", "extractedFields": {{"facultySchool": "CODE or null", "courseCode": "XXXX-XXXX or null", "term": "YYYYST or null", "projectCode": "CODE or null", "subject": "Subject-In-Title-Kebab-Case", "documentForm": "CODE", "date": "{today}", "revision": "Rev0"}}, "reasoning": "Explanation of what I found and why I chose each field.", "confidence": 8}}"""
                }]
            )

        # Parse response
        response_text = message.content[0].text

        # Try to extract JSON from the response
        try:
            # Find JSON in response
            json_match = re.search(r'\{[\s\S]*\}', response_text)
            if json_match:
                result = json.loads(json_match.group())
                return {"success": True, "analysis": result}
        except json.JSONDecodeError:
            pass

        # Fallback if JSON parsing fails
        ext = file_name.split('.')[-1] if '.' in file_name else 'pdf'
        return {
            "success": True,
            "analysis": {
                "suggestedName": f"Document_{today}_Rev0.{ext}",
                "reasoning": "AI analysis completed but response format was unclear",
                "confidence": 5,
                "detectedType": "unknown",
                "suggestedSubject": "Document"
            }
        }

    except anthropic.AuthenticationError:
        return {"success": False, "error": "Invalid API key. Please check your Claude API key."}
    except anthropic.RateLimitError:
        return {"success": False, "error": "Rate limit exceeded. Please wait a moment and try again."}
    except Exception as e:
        return {"success": False, "error": str(e)}


def analyze_with_gemini(api_key: str, content: str, file_name: str, content_type: str, privacy_level: str) -> dict:
    """Analyze file content with Gemini via OpenRouter API.

    Uses google/gemini-2.5-flash model through OpenRouter.
    """
    try:
        today = datetime.now().strftime("%Y-%m-%d")

        # Build the prompt
        base_prompt = f"""You are a file naming expert for UBC CPE (Continuing Professional Education). Analyze this document and suggest a CPE-compliant filename.

Current file: {file_name}
Today's date: {today}

=== STEP 1: CHECK FOR FACULTY-SCHOOL (STRICT MATCHING ONLY) ===
ONLY use Faculty-School codes if you find EXACT matches to these UBC Okanagan names:
- "Irving K. Barber Faculty of Arts and Social Sciences" → IKBASS
- "Irving K. Barber Faculty of Science" → IKBFOS
- "Faculty of Creative and Critical Studies" → FCCS
- "Okanagan School of Education" → OSE
- "Faculty of Applied Science" + "School of Engineering" → APSC-SoE
- "Faculty of Health and Social Development" → FHSD
  - + "School of Nursing" → FHSD-SoN
  - + "School of Social Work" → FHSD-SSW
- "Faculty of Management" → FoM
- "Faculty of Medicine" → MED
- "College of Graduate Studies" → CoGS

⚠️ DO NOT invent Faculty-School codes! If you don't see these EXACT faculty names, set facultySchool to null.
⚠️ "CPE", "ITAL", "55PLUS", course abbreviations, etc. are NOT Faculty-School codes!

=== STEP 2: EXTRACT OTHER FIELDS ===
- COURSE CODE: Four digits + dash + four digits (e.g., 0386-0001). Must be this exact format.
- TERM: Format YYYYST (e.g., 2025WT1). Only if you see clear term/session info.
- PROJECT CODE: Only if explicitly labeled (e.g., "Project: CPE", "Account: XYZ")
- SUBJECT: Use Title-Kebab-Case (e.g., Foundations-For-Restorative-Approach)
  ⚠️ MAX 50 CHARACTERS for subject! Truncate at word boundary if longer.
  Example: "Foundations For A Restorative Approach Health Care Harm And Wellbeing"
  → Truncate to: "Foundations-For-A-Restorative-Approach-Health-Care" (50 chars)
- DOCUMENT FORM: Match keywords to codes:
  LETTERS & CERTIFICATES:
  - "Letter of Proficiency" → LPR | "Letter of Completion" → LCO
  - "Letter of Attendance" → LAT | "Letter of Participation" → LPA

  COMMON TYPES:
  - Report/Count/Chart → RPT | Survey/Analysis → SRY | Summary → SUM
  - Guidelines/Guide → GUI | Manual → MNL | Template → TEM | Schedule → SCH
  - Budget → BGT | Invoice → INV | Contract → CON | Form → FRM
  - Agenda → AGD | Minutes → MIN | Policy → POL | Procedure → PRC
  - Presentation → PRS | Plan → PLN | Proposal → PRO
  - Diagram/Chart → DIA | Screenshot → SCR | Image → IMG

- REVISION: RevA/B/C for drafts, Rev0 for first final, Rev1/2 for subsequent

=== STEP 3: CHOOSE FORMAT ===
COURSE FORMAT (ONLY if you found valid Faculty-School AND course content):
Faculty-School_CourseCode_Term_DocumentForm_Date_Revision.ext

ADVANCED FORMAT (if project code is explicitly present):
ProjectCode_Subject_DocumentForm_Date_Revision.ext

BASIC FORMAT (default - use this for most files):
Subject_DocumentForm_Date_Revision.ext
Example: Enrollment-Count-Report_RPT_{today}_Rev0.xls

=== STEP 4: RESPOND WITH JSON (no markdown, no code blocks) ===
{{"suggestedName": "Filename.ext", "formatUsed": "course|advanced|basic", "extractedFields": {{"facultySchool": "CODE or null", "courseCode": "XXXX-XXXX or null", "term": "YYYYST or null", "projectCode": "CODE or null", "subject": "Subject-In-Title-Kebab-Case", "documentForm": "CODE", "date": "{today}", "revision": "Rev0"}}, "reasoning": "Explanation of what I found and why I chose each field.", "confidence": 8}}"""

        if content_type == "image":
            # Image analysis with Gemini via OpenRouter
            mime_type = content.split(';')[0].replace('data:', '')
            base64_data = content.split(',')[1]

            # Determine if this is likely a PDF (rendered as image) or actual image
            is_pdf = file_name.lower().endswith('.pdf')
            ext = "pdf" if is_pdf else file_name.split('.')[-1] if '.' in file_name else "jpg"

            if is_pdf:
                # PDF rendered as image - Gemini analysis with OCR
                image_prompt = f"""You are a file naming expert for UBC CPE (Continuing Professional Education). READ ALL TEXT in this PDF document image carefully.

Current filename: {file_name}
Today's date: {today}

=== STEP 1: EXTRACT ALL AVAILABLE FIELDS ===

1. FACULTY-SCHOOL (use dash between Faculty and School):
   UBC Okanagan Faculties & Schools - use these abbreviations:
   - "Irving K. Barber Faculty of Arts and Social Sciences" → IKBASS
   - "Irving K. Barber Faculty of Science" → IKBFOS (or IKB-FOS if with school)
   - "Faculty of Creative and Critical Studies" → FCCS
   - "Okanagan School of Education" → OSE
   - "Faculty of Applied Science" / "School of Engineering" → APSC-SoE
   - "Faculty of Health and Social Development" → FHSD
     - with "School of Nursing" → FHSD-SoN
     - with "School of Social Work" → FHSD-SSW
     - with "School of Health and Exercise Sciences" → FHSD-SHES
   - "Faculty of Management" → FoM
   - "Faculty of Medicine" → MED
   - "College of Graduate Studies" → CoGS

   FORMAT: Faculty-School with dash (e.g., FHSD-SoN, APSC-SoE)
   If only Faculty found (no specific school), just use Faculty code (e.g., FCCS)

2. COURSE CODE - Four digits + dash + four digits (e.g., 0386-0001)
   Look for course numbers, section codes

3. TERM OFFERED - Format: YYYYST (Year + Session + Term)
   - Session: W = Winter (Sept-Apr), S = Summer (May-Aug)
   - Term: 1 or 2
   - Sept-Dec → WT1 | Jan-Apr → WT2 | May-Jun → ST1 | Jul-Aug → ST2
   - Example: "December 2025" → 2025WT1
   - Example: "March 2026" → 2025WT2 (still Winter session)

4. PROJECT/ACCOUNT CODE - Project numbers, grant codes (e.g., CPE, PROJ2024)

5. SUBJECT - Use Title-Kebab-Case (e.g., Wildland-Fire-Ecology)
   ⚠️ MAX 50 CHARACTERS! Truncate at word boundary if longer.
   Example: "Foundations-For-A-Restorative-Approach-Health-Care" (50 chars max)

6. DOCUMENT FORM codes:
   LETTERS & CERTIFICATES:
   - "Letter of Proficiency" → LPR | "Letter of Completion" → LCO
   - "Letter of Attendance" → LAT | "Letter of Participation" → LPA
   - "Non-Credit Certificate" → NCC | "MicroCertificate" → NCM

   COMMON TYPES:
   - Agenda → AGD | Agreement → AGR | Budget → BGT | Contract → CON
   - Form → FRM | Guidelines/Guide → GUI | Instructions → INS | Invoice → INV
   - Letter (general) → LTR | Minutes → MIN | Manual → MNL | Plan → PLN
   - Policy → POL | Procedure/SOP → PRC | Proposal → PRO | Presentation → PRS
   - Report → RPT | Schedule → SCH | Template → TEM | Summary → SUM

7. REVISION STATUS:
   - Draft: RevA, RevB, RevC (letters for drafts)
   - Final: Rev0 (first final), Rev1, Rev2 (subsequent finals)
   - Draft after final: Rev0A, Rev0B (number + letter)

=== STEP 2: CHOOSE FORMAT ===

COURSE FORMAT (Faculty + Course content):
Faculty-School_CourseCode_TermOffered_DocumentForm_Date_RevisionStatus.ext
Example: FHSD-SoN_0386-0001_2024WT2_TEM_2025-01-10_Rev0.pptx

ADVANCED FORMAT (Project code found):
ProjectCode_Subject_DocumentForm_Date_RevisionStatus.ext
Example: CPE_Records-Management_POL_2025-01-20_Rev0.pdf

BASIC FORMAT (minimal info):
Subject_DocumentForm_Date_RevisionStatus.ext
Example: Naming-Conventions_LPR_2025-03-11_RevA.docx

=== STEP 3: RESPOND WITH JSON (no markdown, no code blocks) ===
{{"suggestedName": "GeneratedFilename.pdf", "formatUsed": "course|advanced|basic", "extractedFields": {{"facultySchool": "FHSD-SoN or null", "courseCode": "0386-0001 or null", "term": "2024WT2 or null", "projectCode": "CPE or null", "subject": "Subject-In-Title-Kebab-Case", "documentForm": "CODE", "date": "YYYY-MM-DD", "revision": "Rev0"}}, "reasoning": "I found [text]. Faculty: [X]. School: [X]. Term: [X]. Document type: [X].", "confidence": 9}}"""
            else:
                # Actual image file (photo, screenshot, diagram)
                image_prompt = f"""I need help creating a CPE-compliant filename for this image.

IMPORTANT: First, carefully analyze what is shown in this image. Describe what you see in detail.

Current filename: {file_name}

Key elements for the filename:
- Subject (required): What is actually shown in the image. Use Title-Kebab-Case (e.g., Campus-Building). MAX 50 CHARACTERS.
- Document Form: Use IMG for photos, SCR for screenshots, DIA for diagrams
- Date: Use today's date {today} if no date is visible
- Revision: Use 'Rev0' for final version

Respond with ONLY a JSON object (no markdown, no code blocks):
{{"suggestedName": "Content-Description_IMG_{today}_Rev0.{ext}", "reasoning": "I can see [description]. I chose [Subject] because [reason]. I used IMG/SCR/DIA because [reason].", "confidence": 8, "detectedType": "IMG", "suggestedSubject": "Subject-In-Title-Kebab-Case"}}"""

            # OpenRouter request with image
            messages = [{
                "role": "user",
                "content": [
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{mime_type};base64,{base64_data}"
                        }
                    },
                    {
                        "type": "text",
                        "text": image_prompt
                    }
                ]
            }]
        else:
            # Text document analysis - send full content for best analysis
            document_content = content[:4000]  # Limit to 4000 chars for API
            full_prompt = base_prompt + f"\n\nContent: {document_content}"

            messages = [{
                "role": "user",
                "content": full_prompt
            }]

        # Make OpenRouter API request
        response = requests.post(
            OPENROUTER_BASE_URL,
            headers={
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
                "HTTP-Referer": "https://ubc-cpe-naming-tool.streamlit.app",
                "X-Title": "UBC CPE File Naming Tool"
            },
            json={
                "model": GEMINI_MODEL,
                "messages": messages
            },
            timeout=60
        )

        if response.status_code != 200:
            error_data = response.json() if response.text else {}
            error_msg = error_data.get("error", {}).get("message", f"HTTP {response.status_code}")
            return {"success": False, "error": f"API error: {error_msg}"}

        # Parse response
        result_data = response.json()
        response_text = result_data["choices"][0]["message"]["content"]

        # Try to extract JSON from the response
        try:
            # Remove markdown code blocks if present
            cleaned_response = re.sub(r'```json\s*', '', response_text)
            cleaned_response = re.sub(r'```\s*', '', cleaned_response)
            cleaned_response = cleaned_response.strip()

            # Find JSON in response
            json_match = re.search(r'\{[\s\S]*\}', cleaned_response)
            if json_match:
                result = json.loads(json_match.group())
                return {"success": True, "analysis": result}
        except json.JSONDecodeError:
            pass

        # Fallback if JSON parsing fails
        ext = file_name.split('.')[-1] if '.' in file_name else 'pdf'
        return {
            "success": True,
            "analysis": {
                "suggestedName": f"Document_{today}_Rev0.{ext}",
                "reasoning": "AI analysis completed but response format was unclear",
                "confidence": 5,
                "detectedType": "unknown",
                "suggestedSubject": "Document"
            }
        }

    except requests.exceptions.Timeout:
        return {"success": False, "error": "Request timed out. Please try again."}
    except requests.exceptions.RequestException as e:
        return {"success": False, "error": f"Network error: {str(e)}"}
    except Exception as e:
        error_msg = str(e)
        if "429" in error_msg:
            return {"success": False, "error": "Rate limit reached. Please wait a moment and try again."}
        return {"success": False, "error": f"Error: {error_msg[:200]}"}


def generate_filename(format_type: str, subject: str, date_val: datetime, revision: str,
                     extension: str, project_code: str = "", document_form: str = "",
                     faculty_school: str = "", course_code: str = "", term: str = "") -> tuple:
    """Generate CPE-compliant filename. Returns (standard_filename, sharepoint_filename)."""

    formatted_date = date_val.strftime("%Y-%m-%d")
    elements = []

    if format_type == "basic":
        elements = [subject, formatted_date, f"Rev{revision}"]
    elif format_type == "advanced":
        if project_code:
            elements.append(project_code)
        elements.append(subject)
        if document_form:
            elements.append(document_form)
        elements.append(formatted_date)
        elements.append(f"Rev{revision}")
    elif format_type == "course":
        if faculty_school:
            elements.append(faculty_school)
        if course_code:
            elements.append(course_code)
        if term:
            elements.append(term)
        elements.append(subject)
        if document_form:
            elements.append(document_form)
        elements.append(formatted_date)
        elements.append(f"Rev{revision}")

    standard_filename = "_".join(elements) + "." + extension
    sharepoint_filename = " ".join(elements) + "." + extension

    return standard_filename, sharepoint_filename


# Main app header
st.markdown("""
<div class="main-header">
    <h1>UBC CPE File Naming Tool <span class="version-badge">V2 Web</span></h1>
    <p style="color: #666;">Generate standardized filenames and analyze files with AI</p>
</div>
""", unsafe_allow_html=True)

# Main tabs
tab1, tab2 = st.tabs(["📝 Manual Generator", "🤖 AI File Analyzer"])

# ==================== MANUAL GENERATOR TAB ====================
with tab1:
    col1, col2 = st.columns([1, 3])

    with col1:
        st.markdown('<div class="help-panel">', unsafe_allow_html=True)
        st.subheader("Field Guide")

        help_topic = st.selectbox(
            "Select a field to learn more:",
            [""] + list(HELP_CONTENT.keys()),
            format_func=lambda x: HELP_CONTENT[x]["title"] if x else "Select a field..."
        )

        if help_topic:
            st.info(f"**{HELP_CONTENT[help_topic]['title']}**\n\n{HELP_CONTENT[help_topic]['content']}")
        st.markdown('</div>', unsafe_allow_html=True)

    with col2:
        # Format selector
        st.subheader("Choose Naming Format")
        format_type = st.radio(
            "Format",
            ["basic", "advanced", "course"],
            format_func=lambda x: {
                "basic": "Basic Format - For simple documents",
                "advanced": "Advanced Format - For departmental/project documents",
                "course": "Course-Specific Format - For educational materials"
            }[x],
            horizontal=True
        )

        st.divider()

        # Form fields
        col_a, col_b = st.columns(2)

        with col_a:
            subject = st.text_input(
                "Subject/Activity *",
                placeholder="e.g., NamingConventions",
                help="Use PascalCase (capitalize first letter of each word, no spaces)"
            )
            # Auto-format subject
            if subject:
                subject = re.sub(r'\s+', '', subject)
                if subject:
                    subject = subject[0].upper() + subject[1:]

            date_val = st.date_input("Date *", value=datetime.now())

            revision = st.selectbox(
                "Revision Status *",
                list(REVISION_STATUSES.keys()),
                format_func=lambda x: REVISION_STATUSES[x],
                index=3  # Default to "0"
            )

        with col_b:
            extension = st.selectbox(
                "File Extension",
                list(FILE_EXTENSIONS.keys()),
                format_func=lambda x: FILE_EXTENSIONS[x]
            )

            # Advanced fields
            if format_type in ["advanced", "course"]:
                project_code = st.text_input(
                    "Project/Account Number",
                    placeholder="e.g., CPE",
                    help="Optional: Control number, project code, or account identifier"
                )

                document_form = st.selectbox(
                    "Document Form",
                    list(DOCUMENT_FORMS.keys()),
                    format_func=lambda x: DOCUMENT_FORMS[x]
                )
            else:
                project_code = ""
                document_form = ""

            # Course-specific fields
            if format_type == "course":
                faculty_school = st.text_input(
                    "Faculty-School",
                    placeholder="e.g., FHSD-SoN",
                    help="Use dash to separate faculty and school"
                )

                course_code = st.text_input(
                    "Course Code",
                    placeholder="e.g., 0386-0001",
                    help="Four-digit number followed by four-digit section code"
                )

                term = st.text_input(
                    "Term Offered",
                    placeholder="e.g., 2024WT2",
                    help="Format: YYYYST (Year + Session + Term)"
                )
            else:
                faculty_school = ""
                course_code = ""
                term = ""

        # Generate button
        if st.button("Generate Filename", type="primary", use_container_width=True):
            if not subject or not date_val or not revision:
                st.error("Please fill in all required fields (Subject, Date, and Revision Status).")
            else:
                standard_name, sharepoint_name = generate_filename(
                    format_type, subject, date_val, revision, extension,
                    project_code, document_form, faculty_school, course_code, term
                )

                st.markdown('<div class="output-box">', unsafe_allow_html=True)
                st.subheader("Generated Filename")

                # Standard filename
                st.text_input("CPE Standard Filename:", value=standard_name, key="standard_output")
                char_count = len(standard_name)
                if char_count > 150:
                    st.error(f"⚠️ {char_count} characters - Too long! May cause system issues")
                elif char_count > 100:
                    st.warning(f"⚠️ {char_count} characters - Consider shortening")
                else:
                    st.success(f"✓ {char_count} characters")

                # SharePoint filename
                st.text_input("SharePoint Filename (spaces):", value=sharepoint_name, key="sharepoint_output")

                st.markdown('</div>', unsafe_allow_html=True)


# ==================== AI FILE ANALYZER TAB ====================
with tab2:
    col1, col2 = st.columns([2, 1])

    with col2:
        st.subheader("AI Settings")

        # AI Provider selection
        ai_provider = st.radio(
            "Choose AI Provider",
            ["gemini", "claude"],
            format_func=lambda x: {
                "gemini": "🆓 Gemini (FREE)",
                "claude": "💰 Claude (Paid)"
            }[x],
            horizontal=True,
            help="Gemini offers free API access with generous limits"
        )

        if ai_provider == "gemini":
            st.success("Using Gemini 2.5 Flash via OpenRouter")

            # OpenRouter API Key input
            api_key = st.text_input(
                "OpenRouter API Key",
                type="password",
                placeholder="sk-or-v1-...",
                help="Get your free API key at openrouter.ai"
            )

            # Check for API key in secrets (for Streamlit Cloud)
            if not api_key:
                try:
                    api_key = st.secrets.get("OPENROUTER_API_KEY", "")
                except Exception:
                    pass

            st.info("Get your free API key at [openrouter.ai](https://openrouter.ai/keys)")

        else:
            # Claude API Key input
            api_key = st.text_input(
                "Claude API Key",
                type="password",
                placeholder="sk-ant-...",
                help="Your Anthropic API key"
            )

            # Check for API key in secrets (for Streamlit Cloud)
            if not api_key:
                try:
                    api_key = st.secrets.get("ANTHROPIC_API_KEY", "")
                except Exception:
                    pass

            st.info("**Note:** Claude costs approximately $0.005 per file.")

        # Always use full content for best analysis
        privacy_level = "low"

        # How it works
        st.subheader("How it works:")
        provider_name = "Gemini" if ai_provider == "gemini" else "Claude"
        st.markdown(f"""
        1. Upload files below
        2. {provider_name} AI analyzes content
        3. Get CPE-compliant name suggestions
        """)

    with col1:
        st.subheader("Upload Files")

        # File uploader
        uploaded_files = st.file_uploader(
            "Drop files here or click to browse",
            type=['pdf', 'docx', 'doc', 'xlsx', 'xls', 'csv', 'txt', 'jpg', 'jpeg', 'png', 'gif'],
            accept_multiple_files=True,
            help="Supported: PDF, Word, Excel, CSV, TXT, and images"
        )

        if uploaded_files:
            st.write(f"**{len(uploaded_files)} file(s) selected**")

            # Display file list
            for i, file in enumerate(uploaded_files):
                with st.expander(f"📄 {file.name}", expanded=False):
                    st.write(f"Size: {file.size / 1024:.2f} KB")
                    st.write(f"Type: {file.type}")

            # Analyze button
            button_label = f"🤖 Analyze Files with {provider_name} AI" + (" (FREE)" if ai_provider == "gemini" else "")
            if st.button(button_label, type="primary", use_container_width=True):
                if not api_key:
                    if ai_provider == "gemini":
                        st.error("Please enter your OpenRouter API key. Get one free at [openrouter.ai/keys](https://openrouter.ai/keys)")
                    else:
                        st.error("Please enter your Claude API key or configure it in Streamlit secrets.")
                else:
                    progress_bar = st.progress(0)
                    status_text = st.empty()

                    results = []

                    for i, file in enumerate(uploaded_files):
                        status_text.text(f"Analyzing {file.name} with {provider_name}... ({i+1}/{len(uploaded_files)})")
                        progress_bar.progress((i + 1) / len(uploaded_files))

                        # Read file content
                        file.seek(0)  # Reset file pointer
                        content, content_type = read_file_content(file)

                        if content_type == "unknown":
                            results.append({
                                "file": file.name,
                                "success": False,
                                "error": "Unsupported file type"
                            })
                            continue

                        # Analyze with selected AI provider
                        if ai_provider == "gemini":
                            result = analyze_with_gemini(api_key, content, file.name, content_type, privacy_level)
                        else:
                            result = analyze_with_claude(api_key, content, file.name, content_type, privacy_level)
                        result["file"] = file.name
                        results.append(result)

                    status_text.text(f"Analysis complete with {provider_name}!")

                    # Display results
                    st.subheader("Results")

                    for result in results:
                        if result.get("success"):
                            analysis = result["analysis"]

                            # Get format used and extracted fields if available
                            format_used = analysis.get('formatUsed', 'basic')
                            extracted = analysis.get('extractedFields', {})

                            format_label = {
                                'course': '📚 Course Format',
                                'advanced': '📋 Advanced Format',
                                'basic': '📄 Basic Format'
                            }.get(format_used, '📄 Basic Format')

                            st.markdown(f"""
                            <div class="file-suggestion">
                                <h4>📄 {result['file']}</h4>
                                <p><strong>Suggested Name:</strong> <code>{analysis['suggestedName']}</code></p>
                                <p><strong>Format Used:</strong> {format_label}</p>
                            </div>
                            """, unsafe_allow_html=True)

                            # Show extracted fields if available
                            if extracted:
                                with st.expander("📋 Extracted Fields", expanded=True):
                                    cols = st.columns(3)
                                    with cols[0]:
                                        # Check for facultySchool (new format) or faculty (old format)
                                        faculty_val = extracted.get('facultySchool') or extracted.get('faculty')
                                        if faculty_val:
                                            st.markdown(f"**Faculty-School:** {faculty_val}")
                                        if extracted.get('courseCode'):
                                            st.markdown(f"**Course:** {extracted['courseCode']}")
                                        if extracted.get('term'):
                                            st.markdown(f"**Term:** {extracted['term']}")
                                    with cols[1]:
                                        if extracted.get('subject'):
                                            st.markdown(f"**Subject:** {extracted['subject']}")
                                        if extracted.get('documentForm'):
                                            st.markdown(f"**Doc Type:** {extracted['documentForm']}")
                                        if extracted.get('projectCode'):
                                            st.markdown(f"**Project:** {extracted['projectCode']}")
                                    with cols[2]:
                                        if extracted.get('date'):
                                            st.markdown(f"**Date:** {extracted['date']}")
                                        if extracted.get('revision'):
                                            st.markdown(f"**Revision:** {extracted['revision']}")

                            # Show reasoning
                            with st.expander("💭 AI Reasoning"):
                                st.write(analysis.get('reasoning', 'No reasoning provided'))
                                st.write(f"**Confidence:** {analysis.get('confidence', 'N/A')}/10")

                            # Copy button
                            st.text_input(
                                "Copy suggested name:",
                                value=analysis['suggestedName'],
                                key=f"copy_{result['file']}"
                            )

                            st.divider()
                        else:
                            st.error(f"❌ **{result['file']}**: {result.get('error', 'Unknown error')}")
        else:
            st.info("👆 Upload files to get started with AI analysis")


# Footer
st.divider()
st.markdown("""
<div style="text-align: center; color: #666; font-size: 12px;">
    <p>UBC CPE File Naming Tool | Powered by Gemini AI (Free) & Claude AI</p>
    <p>For use by UBC Continuing Professional Education staff</p>
</div>
""", unsafe_allow_html=True)
