"""
File content extraction and processing for the UBC CPE File Naming Tool.

Handles reading content from PDFs, Word documents, Excel files,
images, and plain text files for AI analysis.
"""

import base64
from io import BytesIO

from constants import (
    MAX_TEXT_CONTENT,
    MAX_EXCEL_CONTENT,
    MAX_EXCEL_ROWS,
    MAX_EXCEL_COLS,
    MAX_EXCEL_SHEETS,
    MAX_PDF_PAGES,
    PDF_DPI,
)

# Optional dependency imports with availability flags
try:
    import fitz  # PyMuPDF - for PDF to image conversion (OCR support)
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    from PyPDF2 import PdfReader
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from docx import Document
    DOCX_AVAILABLE = True
except ImportError:
    DOCX_AVAILABLE = False

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False


def pdf_to_image(file_bytes: bytes) -> str | None:
    """Convert first page of PDF to base64 image for AI vision/OCR.

    Returns base64 data URI string for image, or None if conversion fails.
    """
    if not PYMUPDF_AVAILABLE:
        return None

    try:
        pdf_doc = fitz.open(stream=file_bytes, filetype="pdf")
        page = pdf_doc[0]
        mat = fitz.Matrix(PDF_DPI / 72, PDF_DPI / 72)
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        base64_data = base64.b64encode(img_bytes).decode("utf-8")
        pdf_doc.close()
        return f"data:image/png;base64,{base64_data}"
    except Exception:
        return None


def read_pdf_content(file_bytes: bytes) -> str:
    """Extract text from PDF file (fallback when image conversion not available)."""
    if not PDF_AVAILABLE:
        return "PDF parsing not available"

    try:
        pdf_reader = PdfReader(BytesIO(file_bytes))
        text = ""
        for page in pdf_reader.pages[:MAX_PDF_PAGES]:
            text += page.extract_text() or ""
        return text[:MAX_TEXT_CONTENT]
    except Exception as e:
        return f"Could not extract PDF text: {str(e)}"


def read_docx_content(file_bytes: bytes) -> str:
    """Extract text from Word document."""
    if not DOCX_AVAILABLE:
        return "Word document parsing not available"

    try:
        doc = Document(BytesIO(file_bytes))
        text = "\n".join([para.text for para in doc.paragraphs])
        return text[:MAX_TEXT_CONTENT]
    except Exception as e:
        return f"Could not extract Word text: {str(e)}"


def read_xlsx_content(file_bytes: bytes) -> str:
    """Extract text from Excel file with structure preservation."""
    if not XLSX_AVAILABLE:
        return "Excel parsing not available"

    try:
        workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        text = "=== EXCEL FILE CONTENT ===\n"

        for sheet_name in workbook.sheetnames[:MAX_EXCEL_SHEETS]:
            sheet = workbook[sheet_name]
            text += f"\n--- Sheet: {sheet_name} ---\n"

            max_row = min(sheet.max_row or 1, MAX_EXCEL_ROWS)
            max_col = min(sheet.max_column or 1, MAX_EXCEL_COLS)

            row_count = 0
            for row in sheet.iter_rows(
                min_row=1, max_row=max_row, max_col=max_col, values_only=True
            ):
                if any(cell for cell in row):
                    row_text = " | ".join(
                        [str(cell)[:50] if cell else "" for cell in row]
                    )
                    text += row_text.strip() + "\n"
                    row_count += 1
                    if row_count >= 75:
                        text += "... (more rows)\n"
                        break

        return text[:MAX_EXCEL_CONTENT]
    except Exception as e:
        return f"Could not extract Excel text: {str(e)}"


def image_to_base64(file_bytes: bytes, file_name: str) -> str:
    """Convert image file bytes to base64 data URI."""
    ext = file_name.split(".")[-1].lower()
    mime_type = f"image/{ext}" if ext != "jpg" else "image/jpeg"
    base64_data = base64.b64encode(file_bytes).decode("utf-8")
    return f"data:{mime_type};base64,{base64_data}"


def read_file_content(uploaded_file) -> tuple:
    """Read content from uploaded file. Returns (content, content_type).

    For PDFs: Converts to image for AI vision/OCR (much better text detection).
    Falls back to text extraction if image conversion unavailable.
    For images: Returns base64 data URI for vision analysis.
    """
    file_bytes = uploaded_file.read()
    file_name = uploaded_file.name.lower()

    if file_name.endswith(".pdf"):
        pdf_image = pdf_to_image(file_bytes)
        if pdf_image:
            return pdf_image, "image"
        return read_pdf_content(file_bytes), "text"
    elif file_name.endswith(".docx"):
        return read_docx_content(file_bytes), "text"
    elif file_name.endswith((".xlsx", ".xls")):
        return read_xlsx_content(file_bytes), "text"
    elif file_name.endswith((".txt", ".csv")):
        return file_bytes.decode("utf-8", errors="ignore")[:MAX_TEXT_CONTENT], "text"
    elif file_name.endswith((".jpg", ".jpeg", ".png", ".gif")):
        return image_to_base64(file_bytes, file_name), "image"
    else:
        return "Unsupported file type", "unknown"


def compute_file_hash(uploaded_file) -> str:
    """Compute a hash for caching purposes based on file name and size."""
    import hashlib

    uploaded_file.seek(0)
    content = uploaded_file.read()
    uploaded_file.seek(0)
    h = hashlib.md5(content, usedforsecurity=False)
    h.update(uploaded_file.name.encode())
    return h.hexdigest()
